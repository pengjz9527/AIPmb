#!/usr/bin/env python3
"""
生活圈画像分析器
从金融交易数据中推断用户的居住区域、工作区域和生活半径。
"""
import json
import os
from pathlib import Path
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)


# 线上平台关键词（排除）
ONLINE_KEYWORDS = [
    '京东', '淘宝', '唯品会', '抖音', 'QQ', '荣耀', '华为', '支付宝',
    '财付通', '美团平台', '滴滴出行', '哈啰', '亿通行', '轨道', '公交',
    '中铁', '腾讯', '抖店', '抖音月付'
]

# 人名特征（2-4个汉字，无公司/店等后缀）
PERSON_SUFFIXES = ['公司', '店', '医院', '中心', '集团', '银行', '有限', '超市', '商场', '医院']

# 线下实体关键词（保留）
OFFLINE_KEYWORDS = [
    '店', '医院', '中心', '广场', '超市', '市场', '停车场', '物业',
    '小区', '园', '路', '街', '区', '支行'
]

# 商户分类
CATEGORY_RULES = {
    '餐饮美食': ['餐饮', '餐厅', '小吃', '烧烤', '火锅', '面馆', '饭店', '食堂', '美食', '汉堡', '肯德基', '麦当劳', '披萨', '茶姬', '贝甜', '甜品', '咖啡', '肠粉', '肉饼', '春饼', '盖码饭', '湘菜', '川菜', '粤菜'],
    '超市便利店': ['超市', '便利店', '客隆', '7-11', '罗森', '全家', '便利蜂', '美宜佳', '京客隆', '果多美', '稻香村'],
    '生鲜食品': ['盒马', '七鲜', '鲜汇', '生鲜', '蔬菜', '水果'],
    '医疗健康': ['医院', '医疗', '药房', '药店', '诊所', '安贞', '望京医院', '中日友好', '儿研所', '儿科'],
    '商场购物': ['商场', '购物', '万象汇', '华彩', '奥特莱斯', '凯德', '长安商场', '百货'],
    '交通出行': ['地铁', '公交', '停车', '打车', '滴滴', '出租车', '加油', '中石化', '中石油', '一卡通'],
    '生活缴费': ['水电', '燃气', '物业', '供暖', '话费', '联通', '移动'],
}


def is_online_platform(name):
    """判断是否为线上平台"""
    return any(kw in name for kw in ONLINE_KEYWORDS)


def is_person_name(name):
    """判断是否为人名"""
    if len(name) > 4:
        return False
    return not any(s in name for s in PERSON_SUFFIXES)


def categorize_merchant(name):
    """对商户进行分类"""
    for category, keywords in CATEGORY_RULES.items():
        if any(kw in name for kw in keywords):
            return category
    return '其他消费'


def analyze_accounts(filepath):
    """分析账户信息，提取开户行"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    branches = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        branch = row[5]  # 开户行/发卡行
        if branch:
            branches.append(str(branch).strip())
    return branches


def analyze_transactions(filepath):
    """分析交易流水，提取对手方"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    counterparties = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        cp = row[10]  # 对手方名称
        if cp and str(cp) != 'None':
            counterparties.append(str(cp).strip())
    return counterparties


def analyze_credit_card(filepath):
    """分析信用卡账单，提取商户和分类"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    merchants = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant = row[14]  # 商户名称
        if merchant and str(merchant) != 'None':
            merchants.append(str(merchant).strip())
    return merchants


def filter_offline_merchants(merchants):
    """过滤出线下实体商户"""
    offline = []
    for m in merchants:
        if is_online_platform(m):
            continue
        if is_person_name(m):
            continue
        if len(m) <= 2:
            continue
        offline.append(m)
    return offline


def infer_location(branches, offline_merchants):
    """推断居住区域"""
    # 开户行是强信号
    branch_hints = []
    for b in branches:
        if '支行' in b:
            branch_hints.append(b.replace('支行', '').replace('北京', ''))
    
    # 高频商户中的地名
    location_keywords = ['望京', '朝阳', '海淀', '中关村', '西二旗', '国贸', '三里屯', '亚运村', '安贞']
    location_counts = Counter()
    for m in offline_merchants:
        for loc in location_keywords:
            if loc in m:
                location_counts[loc] += 1
    
    return {
        'branches': branch_hints,
        'location_hints': dict(location_counts.most_common(5))
    }


def infer_workplace(counterparties):
    """推断工作区域"""
    # 高频对手方中排除个人和平台
    work_candidates = []
    for cp in counterparties:
        if is_online_platform(cp):
            continue
        if is_person_name(cp):
            continue
        if '公司' in cp or '股份' in cp or '有限' in cp:
            work_candidates.append(cp)
    
    work_counts = Counter(work_candidates)
    top_work = work_counts.most_common(3)
    return [{'name': name, 'count': count} for name, count in top_work]


def analyze_commute(merchants, counterparties):
    """分析通勤方式"""
    all_names = merchants + counterparties
    commute = {
        '地铁': sum(1 for n in all_names if '轨道' in n or '地铁' in n or '亿通行' in n),
        '公交': sum(1 for n in all_names if '公交' in n),
        '打车': sum(1 for n in all_names if '滴滴' in n or '出行' in n),
        '共享单车': sum(1 for n in all_names if '哈啰' in n or '美团单车' in n),
        '开车': sum(1 for n in all_names if '停车' in n or '加油' in n or '中石化' in n or '中石油' in n),
    }
    return commute


def main():
    """主分析函数"""
    base_dir = Path(__file__).parent.parent.parent.parent.parent / 'coredatas'
    
    # 文件路径
    accounts_file = base_dir / '客户账户信息表.xlsx'
    transactions_file = base_dir / '招商银行交易流水汇总明细表.xlsx'
    credit_card_file = base_dir / '信用卡账单合并明细表.xlsx'
    
    result = {
        'summary': {},
        'location_inference': {},
        'workplace_inference': [],
        'offline_merchants_by_category': {},
        'commute_analysis': {},
        'top_offline_merchants': []
    }
    
    # 分析账户信息
    if accounts_file.exists():
        branches = analyze_accounts(accounts_file)
        result['summary']['branches'] = branches
    
    # 分析交易流水
    counterparties = []
    if transactions_file.exists():
        counterparties = analyze_transactions(transactions_file)
    
    # 分析信用卡账单
    merchants = []
    if credit_card_file.exists():
        merchants = analyze_credit_card(credit_card_file)
    
    # 合并所有商户
    all_merchants = merchants + counterparties
    
    # 过滤线下商户
    offline_merchants = filter_offline_merchants(all_merchants)
    offline_counts = Counter(offline_merchants)
    
    # 分类统计
    category_stats = {}
    for merchant, count in offline_counts.most_common(100):
        category = categorize_merchant(merchant)
        if category not in category_stats:
            category_stats[category] = []
        category_stats[category].append({'name': merchant, 'count': count})
    
    result['offline_merchants_by_category'] = category_stats
    result['top_offline_merchants'] = [
        {'name': name, 'count': count}
        for name, count in offline_counts.most_common(30)
    ]
    
    # 推断居住区域
    branches = result.get('summary', {}).get('branches', [])
    result['location_inference'] = infer_location(branches, offline_merchants)
    
    # 推断工作区域
    result['workplace_inference'] = infer_workplace(counterparties)
    
    # 通勤分析
    result['commute_analysis'] = analyze_commute(merchants, counterparties)
    
    # 输出 JSON
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
