#!/usr/bin/env python3
"""
合并信用卡账单PDF和银行交易流水Excel为统一的Excel文件
"""

import os
import re
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

DATA_DIR = '/Users/pengjizhou/Documents/AIPmb/data'
OUTPUT_FILE = os.path.join(DATA_DIR, '信用卡账单合并明细表.xlsx')


def clean_amount(s):
    """清理金额字符串，返回浮点数"""
    s = s.replace(',', '').replace('(CN)', '').strip()
    return float(s)


def resolve_year(trans_month, stmt_year, stmt_month):
    """根据账单年月推算交易日的完整年份"""
    # 交易日月份比账单月份大1以上 → 上一年度的交易
    # 例如：2025年06月账单中出现05月的交易 → 2025年05月
    if trans_month > stmt_month + 1:
        return stmt_year - 1
    # 仅当账单月份是11-12月且交易月份是1-2月时，才推到下一年
    # 例如：2025年12月账单中出现01月的交易 → 2026年01月
    if stmt_month >= 11 and trans_month <= 2:
        return stmt_year + 1
    return stmt_year


def classify_consumption(merchant, category, desc_text):
    """根据商户名称和交易分类推断消费细分子类"""
    m = merchant if merchant else ''
    c = category if category else ''
    d = desc_text if desc_text else ''

    # 非消费类交易
    if c == '还款' or '自动还款' in m:
        return '还款'
    if c == '分期' or '本金' in m:
        return '分期还款'
    if '年费' in m:
        return '年费'
    if c == '退款':
        return '退款'

    # 停车缴费（优先级高，避免被交通出行覆盖）
    parking_kw = ['停车', '停简单', 'ETCP', '驿停车', 'P云', '互联互通停车',
                  '阳光海天', '聚宝充']
    if any(kw in m for kw in parking_kw):
        return '停车缴费'

    # 网约车
    if '滴滴' in m or '曹操' in m or 'T3出行' in m:
        return '网约车'

    # 公共交通（地铁、公交）
    transit_kw = ['地铁', '轨道交通', '公交', '公共交通', '亿通行', '广州骑安',
                  '广州地铁', '南京地铁', '南昌地铁']
    if any(kw in m for kw in transit_kw):
        return '公共交通'

    # 铁路出行
    if '中铁网络' in m or '中铁' in m and '便利' not in m and '快运' not in m:
        return '铁路出行'

    # 过路/收费站
    toll_kw = ['收费站', '机场路投资', '江西省交通监控']
    if any(kw in m for kw in toll_kw):
        return '过路收费'

    # 共享单车/助力车
    shared_kw = ['哈啰', '助力车', '骑安']
    if any(kw in m for kw in shared_kw):
        return '共享单车'

    # 加油
    fuel_kw = ['加油', '中石化', '中石油', '加油站']
    if any(kw in m for kw in fuel_kw):
        return '加油'

    # 快递物流
    logistics_kw = ['顺丰', '快运', '快递', '中铁快运']
    if any(kw in m for kw in logistics_kw):
        return '快递物流'

    # 超市便利店
    supermarket_kw = ['超市', '便利店', '盒马', '七鲜', '罗森', '美宜佳',
                      '7-11', '沃尔玛', '京客隆', '永辉', '鲜汇生活',
                      '中铁便利店', '果多美', '果栗园', '嗨特购', 'HitGoo',
                      '柒一拾壹']
    if any(kw in m for kw in supermarket_kw):
        return '超市便利店'

    # 生鲜食品
    fresh_kw = ['生鲜', '多点新鲜', '象鲜', '蔬菜种子', '天北生鲜',
                '黑土地自产', '杨博士生态散养']
    if any(kw in m for kw in fresh_kw):
        return '生鲜食品'

    # 餐饮强标识优先（防止"护国寺小吃北京安贞店"被误判为医疗）
    dining_strong = ['小吃', '餐厅', '美食', '火锅', '烧烤', '面馆', '炸鸡',
                     '汉堡', '肯德基', '麦当劳', '金拱门', '包子', '奶茶',
                     '咖啡', '甜品', '巴黎贝甜', '稻香村', '聚宝源', '紫光园',
                     '眉州东坡', '护国寺', '鸿兴美食', '嘉和一品', '老乡鸡',
                     '德克士', '南城香', '真功夫', '呷哺', '煲仔', '烤串',
                     '面', '粉', '饺', '粥', '饭', '串', '烫', '卤',
                     'FOODBOWL', '超级碗', '美团平台商户']
    if any(kw in m for kw in dining_strong):
        return '餐饮美食'

    # 医疗健康
    medical_kw = ['医院', '医药', '药房', '药店', '药', '安贞', '协和',
                  '儿科', '儿童医院', '正济堂', '111医药']
    if any(kw in m for kw in medical_kw):
        return '医疗健康'

    # 旅游住宿
    travel_kw = ['携程', '酒店', '金陵状元楼', '旅游', '景区', '鹏程万里']
    if any(kw in m for kw in travel_kw):
        return '旅游住宿'

    # 保险
    if '保险' in m:
        return '保险'

    # 教育学习
    edu_kw = ['学校', '新东方', '学堂', '考试', '教育']
    if any(kw in m for kw in edu_kw):
        return '教育学习'

    # 通讯/软件/数码
    digital_kw = ['华为', '腾讯', 'QQ阅读', '百度网盘', '联通', '月之暗面',
                  '荣耀', '软件', '高德', '高德地图', '甲虎科技', 'App',
                  '拉扎斯', '萤启', '速通科技', '新北洋']
    if any(kw in m for kw in digital_kw):
        return '通讯软件数码'

    # 网购电商
    ecommerce_kw = ['京东', '唯品会', '小米', '淘宝', '宜家', '老饭骨官方旗舰',
                    '卢正浩', '华宝在线', '英斯迪尔旗舰', '石头科技',
                    '魔声', 'Apple产品', 'KEKLLE', '惠普打印机',
                    '手机自营旗舰', '松山棉店京东自营', '蔬菜源头直发']
    if any(kw in m for kw in ecommerce_kw):
        return '网购电商'

    # 餐饮美食（关键词最多，放在超市后面避免误匹配）
    dining_kw = ['餐饮', '美食', '餐厅', '小吃', '烧烤', '火锅', '面馆',
                 '炸鸡', '汉堡王', '肯德基', '麦当劳', '金拱门', '粥',
                 '包子', '奶茶', '咖啡', '甜品', '巴黎贝甜', '稻香村',
                 '美团平台商户', '烤肉', '烤串', '小吃店', '食品',
                 '饭', '菜', '厨', '鸭货', '鸡', '羊肉', '牛肉',
                 '茶餐厅', '料理', '煲仔', '串', '煎', '炒',
                 '卤', '面', '粉', '饺', '锅贴', '馄饨',
                 '烧鸡', '烧肉', '烤鸭', '烧鹅', '麻辣', '酸辣',
                 '嘉和一品', '老乡鸡', '德克士', '南城香', '真功夫',
                 '呷哺', '必胜客', '披萨', '寿司', '拉面', '盖码',
                 '煲', '烫', '味', '香', '甜', '酸', '辣',
                 '聚宝源', '紫光园', '眉州东坡', '护国寺', '糖葫芦',
                 '百年義利', '早餐', '万家早安', '坐门墩儿',
                 '新田餐饮', '同好餐饮', '喆喆餐饮', '鸿兴美食',
                 '西部马华', '小骆驼', '庆丰', '长安商场',
                 '马小火', '美栗乡', '沪上阿姨',
                 '霸王茶姬', 'CHAGEE', '紫气羊来',
                 '解药咖啡', '肥猫茶', '蔡林记',
                 '赏啫就啫', '农耕故事', '正宗保山',
                 '海口龙华', '海南星图', '盖东来', '椒野香',
                 '拾陆两自助', '小锅小灶', '江西小炒', '香一村',
                 '李与白', '千忆汤包', '一围肥牛', '怡海优食',
                 '天津胜津', '熊吞', '正好',
                 '漠上羊', '三千里音乐', '清真楼兰', '清真',
                 '德同轩', '超意兴', '武哥春饼', '湘小盘',
                 '兰湘子', '禾子炸鸡', '面郎哥', '重庆小面',
                 '得意美食', '爱汤坊', '香宝宝', '延吉市丰茂',
                 '郭记糖葫芦', '糖亿佳', '巍云小吃',
                 '幸润食品', '润田食品',
                 'FOODBOWL', '超级碗', '左庭右院',
                 '萍姐', '煲仔哥', '云味集',
                 '贵州丁家', '金牛区老倪', '四川四娃娃',
                 '济南华源', '铜仁市碧江', '宁夏盐池',
                 '安徽美雉', '涪陵区', '金堂县五凤',
                 '北流市梅兴', '四川小隐',
                 '四川新派', '四川古蜀', '广州新渝城',
                 '湖北谦诚', '开封市祥符',
                 '隆尧至诚', '河曲县烜宴', '长沙市天心',
                 '江宁区宏之景', '麻城市满路花',
                 '山海关古文化', '鱼酒鲜儿',
                 '北京水西庄', '广东道至正', '这招很红',
                 '广东道2店', '北京一轩珍味',
                 '北京真功夫',
                 '北京一手店',
                 '廊坊致匠', '黑龙江采信',
                 '重庆金和友', '重庆诚企瑞',
                 '上海瑞象',
                 '甘星辰', '陈文宪', '宋相旗', '周宏渚',
                 '商户杨于战', '商户_印世国',
                 '大理市云姝', '大理市正宗',
                 '福州金拱门', '郑州仟吉', '10－特产',
                 '北京市明洞邦',
                 '同仁药食',
                 '三快在线',
                 ]
    # 美团App团购、美团App开头的商户
    if '美团App' in m:
        return '餐饮美食'
    if any(kw in m for kw in dining_kw):
        return '餐饮美食'
    # 美团平台商户默认归为餐饮
    if m == '美团平台商户':
        return '餐饮美食'

    # 日用百货
    daily_kw = ['名创优品', '九木杂物社', '万象汇', '凯德MALL', '凯德',
                '合生麒麟', '商场', '百货', '奥特莱斯', '赛特',
                '凯城时代', '龙湖', '物业管理', '物业',
                '首开望京', '三河航康', '航星园', 'i友未来',
                '北京清河万象', '松山（北京',
                ]
    if any(kw in m for kw in daily_kw):
        return '日用百货'

    # 服装鞋帽
    clothing_kw = ['优衣库', '李宁', '百丽', '服装', '鞋', '棉店',
                   '时装', '快尚', '双桥区莹雪']
    if any(kw in m for kw in clothing_kw):
        return '服装鞋帽'

    # 物业/房产
    property_kw = ['物业', '首开', '三河航康', '航星园', 'i友未来',
                   '龙湖智创', '北京魏公村综合体']
    if any(kw in m for kw in property_kw):
        return '物业房产'

    # 生活缴费
    utility_kw = ['缴费', '水电', '燃气', '供暖', '热力',
                  '海淀区城市管理', '非税']
    if any(kw in m for kw in utility_kw):
        return '生活缴费'

    # 文化娱乐
    entertain_kw = ['电影', '阅读', 'KTV', '景区', '文化馆', '正好文化']
    if any(kw in m for kw in entertain_kw):
        return '文化娱乐'

    # 充值缴费（话费等）
    if '充值' in m or '话费' in m:
        return '充值缴费'

    # 通用的支付平台交易
    generic_pay = ['财付通', '支付宝-消费', '支付宝-支付宝-消费',
                   '美团支付', '美团', '扫二维码付款']
    if m in generic_pay:
        return '其他消费'

    # 兜底
    return '其他消费'


def get_stmt_info(filename):
    """从文件名提取账单年月"""
    m = re.search(r'(\d{4})年(\d{2})月', filename)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def parse_credit_card_pdf(filepath):
    """解析信用卡账单PDF，提取交易记录"""
    filename = os.path.basename(filepath)
    stmt_year, stmt_month = get_stmt_info(filename)
    stmt_period = f"{stmt_year}年{stmt_month:02d}月"

    doc = fitz.open(filepath)
    records = []
    category = ''

    date_re = re.compile(r'^(\d{2})/(\d{2})$')
    num_re = re.compile(r'^-?[\d,]+(?:\.\d+)?(?:\(CN\))?$')

    for page in doc:
        text = page.get_text("text")
        lines = text.split('\n')

        i = 0
        while i < len(lines):
            line = lines[i].strip()

            # 检测分类标记
            if line in ('还款', '退款', '消费', '分期'):
                category = line
                i += 1
                continue

            m = date_re.match(line)
            if not m:
                i += 1
                continue

            # 阶段1：收集日期（1-2个日期行）
            items = []
            while i < len(lines):
                l = lines[i].strip()
                if not l:
                    i += 1
                    continue
                if date_re.match(l):
                    items.append(('date', l))
                    i += 1
                    if len(items) >= 2:
                        break
                else:
                    break

            # 阶段2：收集描述文本（遇到数字则停止）
            while i < len(lines):
                l = lines[i].strip()
                if not l:
                    i += 1
                    continue
                if num_re.match(l):
                    break
                if l in ('还款', '退款', '消费', '分期'):
                    break
                if date_re.match(l):
                    break
                if l.startswith('招商银行信用卡') or l.startswith('CMB Credit'):
                    break
                if l.startswith('本期应还') or l.startswith('上期账单'):
                    break
                if l.startswith('(1)') or l.startswith('(2)') or l.startswith('(3)'):
                    break
                if l.startswith('★') or l.startswith('人民币账户') or l.startswith('交易日'):
                    break
                items.append(('desc', l))
                i += 1

            # 阶段3：收集恰好3个数字字段
            while i < len(lines) and len([x for x in items if x[0] == 'num']) < 3:
                l = lines[i].strip()
                if not l:
                    i += 1
                    continue
                if num_re.match(l):
                    items.append(('num', l))
                    i += 1
                else:
                    break

            # 提取数字字段
            nums = [v for t, v in items if t == 'num']

            if len(nums) < 3:
                continue

            # 分离日期和描述
            dates = [v for t, v in items if t == 'date']
            desc_parts = [v for t, v in items if t == 'desc']

            trans_date = ''
            post_date = ''
            if len(dates) >= 2:
                trans_date = dates[0]
                post_date = dates[1]
            elif len(dates) == 1:
                trans_date = dates[0]
                post_date = dates[0]

            desc_text = ''.join(desc_parts).strip()

            amount = clean_amount(nums[0])
            card_digits = nums[1]
            original_amount_str = nums[2]
            original_amount = clean_amount(original_amount_str)

            direction = '收入' if amount < 0 else '支出'

            # 拆分交易摘要为支付渠道和商户名称
            payment_channel = ''
            merchant = desc_text
            if '-' in desc_text:
                parts = desc_text.split('-', 1)
                payment_channel = parts[0]
                merchant = parts[1]

            # 推算完整日期（统一为 YYYY-MM-DD 格式）
            trans_mm = int(trans_date.split('/')[0])
            trans_dd = int(trans_date.split('/')[1])
            full_trans_year = resolve_year(trans_mm, stmt_year, stmt_month)
            full_trans_date = f"{full_trans_year}-{trans_mm:02d}-{trans_dd:02d}"

            post_mm = int(post_date.split('/')[0])
            post_dd = int(post_date.split('/')[1])
            full_post_year = resolve_year(post_mm, stmt_year, stmt_month)
            full_post_date = f"{full_post_year}-{post_mm:02d}-{post_dd:02d}"

            record = {
                '数据来源': '信用卡账单',
                '账单月份': stmt_period,
                '年份': stmt_year,
                '月份': stmt_month,
                '持卡人姓名': '彭楫洲',
                '卡号/账号': f'****{card_digits}',
                '开户行': '',
                '交易日期': full_trans_date,
                '记账日期': full_post_date,
                '交易年份': full_trans_year,
                '交易月份': trans_mm,
                '交易分类': category,
                '收支方向': direction,
                '交易摘要': desc_text,
                '支付渠道': payment_channel,
                '商户名称': merchant,
                '消费细分子类': classify_consumption(merchant, category, desc_text),
                '交易金额': abs(amount),
                '交易地金额': abs(original_amount),
                '联机余额': '',
                '对手方名称': '',
                '对手方账号': '',
                '币种': 'CNY',
            }
            records.append(record)

    doc.close()
    return records


def read_debit_card_excel(filepath):
    """读取银行交易流水汇总明细表Excel"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['交易流水明细']

    records = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row[0] is None or not isinstance(row[0], (int, float)):
            continue

        direction = str(row[5]) if row[5] else ''
        amount = float(row[6]) if row[6] else 0

        # 交易摘要(原始) = 交易大类
        trans_category = str(row[8]) if row[8] else ''
        # 交易方式/渠道
        trans_method = str(row[9]) if row[9] else ''
        # 对手方名称
        counter_name = str(row[10]) if row[10] else ''
        # 对手方账号
        counter_account = str(row[11]) if row[11] else ''
        # 来源账号
        source_account = str(row[12]) if row[12] else ''
        # 开户行
        branch = str(row[13]) if row[13] else ''

        # 构建交易摘要
        if counter_name and trans_method:
            desc = f"{trans_method} - {counter_name}"
        elif counter_name:
            desc = counter_name
        elif trans_method:
            desc = trans_method
        else:
            desc = trans_category

        # 商户名称：优先对手方名称
        merchant = counter_name if counter_name else (trans_method if trans_method else trans_category)

        # 交易日期处理
        date_str = str(row[1]) if row[1] else ''
        trans_year = int(row[2]) if row[2] else 0
        trans_month = int(row[3]) if row[3] else 0

        record = {
            '数据来源': '借记卡流水',
            '账单月份': f"{trans_year}年{trans_month:02d}月" if trans_year else '',
            '年份': trans_year,
            '月份': trans_month,
            '持卡人姓名': '彭楫洲',
            '卡号/账号': source_account,
            '开户行': branch,
            '交易日期': date_str,
            '记账日期': date_str,
            '交易年份': trans_year,
            '交易月份': trans_month,
            '交易分类': trans_category,
            '收支方向': direction,
            '交易摘要': desc,
            '支付渠道': trans_method,
            '商户名称': merchant,
            '交易金额': amount,
            '交易地金额': amount,
            '联机余额': float(row[7]) if row[7] else '',
            '对手方名称': counter_name,
            '对手方账号': counter_account,
            '币种': str(row[4]) if row[4] else 'CNY',
        }
        records.append(record)

    wb.close()
    return records


def write_output_excel(records, output_path):
    """写入合并后的Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '交易流水合并明细'

    headers = [
        '序号', '账单月份', '年份', '月份',
        '持卡人姓名', '卡号',
        '交易日期', '记账日期', '交易年份', '交易月份',
        '交易分类', '收支方向',
        '交易摘要', '支付渠道', '商户名称', '消费细分子类',
        '交易金额', '交易地金额',
        '币种',
    ]

    col_widths = [8, 14, 8, 8, 12, 12, 14, 14, 8, 8, 14, 10, 40, 16, 35, 14, 14, 14, 8]

    # 样式定义
    header_font = Font(name='微软雅黑', bold=True, size=10)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    data_font = Font(name='微软雅黑', size=9)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    number_format_amt = '#,##0.00'

    # 设置列宽
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 写入数据行
    for idx, record in enumerate(records, 1):
        row = idx + 1

        values = [
            idx,
            record['账单月份'],
            record['年份'],
            record['月份'],
            record['持卡人姓名'],
            record['卡号/账号'],
            record['交易日期'],
            record['记账日期'],
            record['交易年份'],
            record['交易月份'],
            record['交易分类'],
            record['收支方向'],
            record['交易摘要'],
            record['支付渠道'],
            record['商户名称'],
            record['消费细分子类'],
            record['交易金额'],
            record['交易地金额'],
            record['币种'],
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.border = thin_border

            # 金额列特殊格式（第17、18列）
            if col in (17, 18) and isinstance(value, (int, float)):
                cell.number_format = number_format_amt
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col in (1, 3, 4, 9, 10):
                cell.alignment = center_align
            elif col == 13:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # 添加筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records)+1}"

    wb.save(output_path)
    print(f"已保存: {output_path}")
    print(f"共 {len(records)} 条记录")


def main():
    # 1. 解析所有信用卡账单PDF
    credit_records = []
    pdf_files = sorted([
        f for f in os.listdir(DATA_DIR)
        if '信用卡账单' in f and f.endswith('.pdf')
    ])
    print(f"发现 {len(pdf_files)} 个信用卡账单PDF文件")

    for f in pdf_files:
        filepath = os.path.join(DATA_DIR, f)
        try:
            records = parse_credit_card_pdf(filepath)
            credit_records.extend(records)
            print(f"  {f}: 提取 {len(records)} 条记录")
        except Exception as e:
            print(f"  {f}: 解析失败 - {e}")

    print(f"信用卡账单共提取 {len(credit_records)} 条记录")

    # 2. 按交易日期排序
    all_records = credit_records

    def sort_key(r):
        date_str = r.get('交易日期', '')
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        except (ValueError, TypeError):
            date_obj = datetime.max
        return date_obj

    all_records.sort(key=sort_key)

    print(f"\n共 {len(all_records)} 条记录，正在生成Excel...")

    # 3. 写入Excel
    write_output_excel(all_records, OUTPUT_FILE)

    # 打印统计信息
    income = sum(r['交易金额'] for r in all_records if r['收支方向'] == '收入')
    expense = sum(r['交易金额'] for r in all_records if r['收支方向'] == '支出')
    print(f"\n统计:")
    print(f"  总记录数: {len(all_records)} 条")
    print(f"  总收入(退款/还款): {income:,.2f} 元")
    print(f"  总支出(消费): {expense:,.2f} 元")
    if all_records:
        first_date = all_records[0]['交易日期']
        last_date = all_records[-1]['交易日期']
        print(f"  时间范围: {first_date} ~ {last_date}")


if __name__ == '__main__':
    main()
