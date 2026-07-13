from functools import lru_cache
from pathlib import Path

import openpyxl

from pmb.core.config import (
    ACCOUNT_FILE,
    CREDIT_TX_FILE,
    DEBIT_TX_FILE,
    PAYMENT_FILE,
    PRODUCT_FILES,
    HELD_PRODUCT_FILES,
    HEADER_CONFIG,
)


def _load_excel(filepath: Path, header_row: int, data_start: int, footer_skip: int = 0) -> list[dict]:
    """通用Excel加载器，返回 list[dict]"""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    # 读取表头
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        headers.append(str(v) if v is not None else f"_col{c}")

    # 读取数据行
    rows = []
    total_rows = ws.max_row - footer_skip
    for r in range(data_start, total_rows + 1):
        row_dict = {}
        all_none = True
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            if v is not None:
                all_none = False
            row_dict[h] = v
        if not all_none:
            rows.append(row_dict)

    wb.close()
    return rows


class DataLoader:
    """数据加载器，使用缓存避免重复读取"""

    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self):
        if not hasattr(self, "_initialized"):
            self._initialized = True
            self._cache = {}

    def load_accounts(self) -> list[dict]:
        if "accounts" not in self._cache:
            hr, ds, fs = HEADER_CONFIG["account"]
            self._cache["accounts"] = _load_excel(ACCOUNT_FILE, hr, ds, fs)
        return self._cache["accounts"]

    def load_credit_transactions(self) -> list[dict]:
        if "credit_tx" not in self._cache:
            hr, ds, fs = HEADER_CONFIG["credit_tx"]
            self._cache["credit_tx"] = _load_excel(CREDIT_TX_FILE, hr, ds, fs)
        return self._cache["credit_tx"]

    def load_debit_transactions(self) -> list[dict]:
        if "debit_tx" not in self._cache:
            hr, ds, fs = HEADER_CONFIG["debit_tx"]
            self._cache["debit_tx"] = _load_excel(DEBIT_TX_FILE, hr, ds, fs)
        return self._cache["debit_tx"]

    def load_products(self, category: str) -> list[dict]:
        key = f"product_{category}"
        if key not in self._cache:
            filepath = PRODUCT_FILES.get(category)
            if not filepath or not filepath.exists():
                return []
            hr, ds, fs = HEADER_CONFIG["product"]
            self._cache[key] = _load_excel(filepath, hr, ds, fs)
        return self._cache[key]

    def load_all_products(self) -> dict[str, list[dict]]:
        result = {}
        for cat in PRODUCT_FILES:
            result[cat] = self.load_products(cat)
        return result

    def load_held_wealth(self) -> list[dict]:
        if "held_wealth" not in self._cache:
            hr, ds, fs = HEADER_CONFIG["held_wealth"]
            self._cache["held_wealth"] = _load_excel(HELD_PRODUCT_FILES["wealth"], hr, ds, fs)
        return self._cache["held_wealth"]

    def load_held_loans(self) -> list[dict]:
        if "held_loans" not in self._cache:
            hr, ds, fs = HEADER_CONFIG["held_loan"]
            self._cache["held_loans"] = _load_excel(HELD_PRODUCT_FILES["loan"], hr, ds, fs)
        return self._cache["held_loans"]

    def load_held_pensions(self) -> list[dict]:
        if "held_pensions" not in self._cache:
            hr, ds, fs = HEADER_CONFIG["held_pension"]
            self._cache["held_pensions"] = _load_excel(HELD_PRODUCT_FILES["pension"], hr, ds, fs)
        return self._cache["held_pensions"]

    def load_payments(self) -> list[dict]:
        """加载缴费记录"""
        if "payments" not in self._cache:
            hr, ds, fs = HEADER_CONFIG["payment"]
            self._cache["payments"] = _load_excel(PAYMENT_FILE, hr, ds, fs)
        return self._cache["payments"]

    def reload(self):
        """清除缓存"""
        self._cache.clear()


# 全局单例
loader = DataLoader()
