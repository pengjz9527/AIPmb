"""收入趋势预测与消费规划 Skill — 分析收入趋势、预测未来收支、给出消费调整策略

关键优化：
- 剔除个人间转账（行内转账转入/网联收款/转账等），不计入真实收入
- 单独标记个人转入资金，推断用途（贷款还款、大额消费、账户调拨等）
"""
import json
from collections import defaultdict, Counter
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

from pmb.skills.base import BaseSkill, SkillResult
from pmb.skills.api_client import create_client


# ---- 个人转账识别规则 ----

# 交易大类：明确属于个人间资金划转，非外部收入
PERSONAL_TRANSFER_CATEGORIES = {
    '转账', '行内转账转入', '网联收款', '汇入汇款',
}

# 人名排除后缀（含这些关键词即为机构，非个人）
PERSON_NAME_EXCLUDE_SUFFIXES = [
    '公司', '银行', '有限', '支行', '集团', '中心', '医院',
    '学校', '超市', '商场', '餐厅', '酒店', '科技', '网络',
    '分行', '营业部', '分理处', '储蓄所', '合作社', '政府',
    '局', '委', '办', '站', '所',
]

# 后续消费推断用途的规则
TRANSFER_PURPOSE_RULES = [
    # (关键词列表, 用途标签, 描述模板)
    (['贷款', '还贷', '房贷', '月供', '还款', '按揭', '公积金'], 'loan_repayment', '转入用于偿还贷款'),
    (['理财', '基金', '存款', '投资', '黄金', '保险', '国债'], 'investment', '转入用于投资理财'),
    (['购房', '买车', '首付', '定金', '装修'], 'major_purchase', '转入用于大额消费'),
    (['学费', '培训', '教育'], 'education', '转入用于教育培训'),
    (['医疗', '住院', '手术', '药'], 'medical', '转入用于医疗支出'),
]


def _is_person_name(name: str) -> bool:
    """判断对手方名称是否为个人姓名（2-4个汉字，不含机构后缀）"""
    if not name or len(name) < 2 or len(name) > 4:
        return False
    # 排除占位符
    if name in ("None", "未知", "null", ""):
        return False
    # 必须全部是中文汉字（排除英文字母、数字、特殊符号）
    if not all('\u4e00' <= ch <= '\u9fff' for ch in name):
        return False
    return not any(s in name for s in PERSON_NAME_EXCLUDE_SUFFIXES)


def _is_personal_transfer(category: str, counterparty: str) -> bool:
    """判断一条收入记录是否为个人间转账（非真实收入）"""
    # 交易大类必须是转账类
    if category not in PERSONAL_TRANSFER_CATEGORIES:
        return False
    # 对手方是个人姓名（非机构）
    return _is_person_name(counterparty)


class IncomeForecastSkill(BaseSkill):
    """收入趋势预测与消费规划：
    1. 分析历史收入趋势（工资稳定性、奖金规律、收入波动）
    2. 分析消费支出习惯（刚性/弹性支出、消费频次变化）
    3. 预测未来收支平衡（收入能否覆盖消费增长）
    4. 给出消费调整策略（量入为出、优化建议）
    """

    @property
    def name(self) -> str:
        return "income_forecast"

    @property
    def description(self) -> str:
        return (
            "预测我的未来收入趋势和消费支出，评估收入能否覆盖未来消费增长，"
            "并给出消费调整策略。适用场景：'我未来的收入怎么样'、'收入能不能覆盖支出'、"
            "'消费趋势预测'、'量入为出规划'、'收支平衡分析'、'消费调整建议'等。"
        )

    @property
    def parameters_schema(self) -> dict:
        return {"type": "object", "properties": {}}

    async def execute(self, **kwargs) -> SkillResult:
        user_name = kwargs.get("user_name", "")
        api = create_client(user_name=user_name)

        # 1. 从交易流水和信用卡账单提取数据
        raw_data = self._analyze_raw_data(user_name=user_name)

        # 2. 从 API 获取消费统计
        monthly_consumption, _ = await api.get_consumption_stats(group_by="month", top=24)
        category_consumption, _ = await api.get_consumption_stats(group_by="subcategory", top=15)

        # 3. 获取账户余额
        summary_data, _ = await api.get_account_summary()
        summary = {item["label"]: item["value"] for item in summary_data} if summary_data else {}
        all_accounts, _ = await api.list_accounts(limit=100)
        total_balance = 0.0
        for a in all_accounts:
            at = str(a.get("account_type", ""))
            if "借记" in at:
                try:
                    total_balance += float(a.get("balance", 0) or 0)
                except (TypeError, ValueError):
                    pass

        # 4. 计算关键指标 —— 使用「真实收入」（已剔除个人转入）
        real_income = raw_data.get("real_monthly_income", {})
        # 如果 real_monthly_income 为空（比如 Excel 未加载），回退到 monthly_income
        income_for_trend = real_income if real_income else raw_data.get("monthly_income", {})
        income_trend = self._compute_income_trend(income_for_trend)
        expense_trend = self._compute_expense_trend(raw_data.get("monthly_expense", {}))
        forecast = self._forecast_balance(income_trend, expense_trend)
        strategies = self._generate_strategies(income_trend, expense_trend, forecast, total_balance)

        # 5. 构建转入资金警告
        transfer_summary = raw_data.get("transfer_summary", {})
        transfer_warning = ""
        if transfer_summary.get("total_count", 0) > 0:
            total_transfer = transfer_summary.get("total_amount", 0)
            if income_for_trend:
                total_inc = sum(income_for_trend.values())
                transfer_pct = total_transfer / max(total_inc, 1) * 100
                transfer_warning = (
                    f"注意：共{transfer_summary['total_count']}笔个人转入资金合计¥{total_transfer:,.0f}，"
                    f"占记录总收入的{transfer_pct:.0f}%，已从收入趋势分析中剔除"
                )
            else:
                transfer_warning = (
                    f"注意：共{transfer_summary['total_count']}笔个人转入资金合计¥{total_transfer:,.0f}，"
                    f"已从收入分析中剔除"
                )

        summary_parts = [
            f"已分析{income_trend.get('total_months', 0)}个月收支数据",
            f"识别{len(raw_data.get('income_sources', []))}类收入来源（已剔除{transfer_summary.get('total_count', 0)}笔个人转入）",
            f"{len(raw_data.get('expense_categories', []))}类支出方向",
            f"预测未来{forecast.get('forecast_years', 2)}年收支趋势",
        ]

        return SkillResult(
            success=True,
            data={
                "income_trend": income_trend,
                "expense_trend": expense_trend,
                "credit_card_trend": raw_data.get("credit_card_trend", {}),
                "income_sources": raw_data.get("income_sources", []),
                "expense_categories": raw_data.get("expense_categories", []),
                "anomalies": raw_data.get("anomalies", []),
                "monthly_consumption": monthly_consumption,
                "category_consumption": category_consumption,
                "account_balance": round(total_balance, 2),
                "account_summary": summary,
                "forecast": forecast,
                "strategies": strategies,
                # 新增：个人转入相关数据
                "transfer_summary": transfer_summary,
                "transfer_inference": raw_data.get("transfer_inference", {}),
                "transfer_warning": transfer_warning,
                # 保留原始总收入（含转入）供参考
                "all_monthly_income": raw_data.get("monthly_income", {}),
            },
            summary="；".join(summary_parts),
        )

    def _analyze_raw_data(self, user_name: str = "") -> dict:
        """解析交易流水和信用卡账单原始数据

        区分「真实收入」与「个人转入」：
        - 真实收入：代发工资、报销款、公积金、利息、基金赎回等
        - 个人转入：行内转账转入/网联收款/转账等，来自个人账户的资金划转
        """
        base_dir = Path(__file__).resolve().parent.parent.parent.parent / "coredatas"
        transactions_file = base_dir / "银行交易流水汇总明细表.xlsx"
        credit_card_file = base_dir / "信用卡账单合并明细表.xlsx"

        result = {
            "monthly_income": {},
            "monthly_expense": {},
            "income_sources": [],
            "expense_categories": [],
            "credit_card_trend": {},
            "anomalies": [],
            # 新增：个人转入相关字段
            "real_monthly_income": {},          # 剔除个人转入后的真实月收入
            "personal_transfers": [],           # 个人转入明细
            "transfer_inference": {},           # 转入资金用途推断
            "transfer_summary": {},             # 转入汇总
        }

        if openpyxl is None:
            return result

        # 解析交易流水
        if transactions_file.exists():
            wb = openpyxl.load_workbook(transactions_file)
            ws = wb.active

            monthly_income = defaultdict(float)
            real_monthly_income = defaultdict(float)  # 剔除转入后的真实收入
            monthly_expense = defaultdict(float)
            income_sources = Counter()
            expense_categories = Counter()
            monthly_income_detail = defaultdict(list)

            # 收集所有支出记录（用于推断转入用途）
            all_expenses = []
            # 个人转入明细
            personal_transfers = []

            for row in ws.iter_rows(min_row=4, values_only=True):
                try:
                    year = int(row[3])
                    month = int(row[4])
                except (ValueError, TypeError):
                    continue

                # 用户姓名过滤（第1列是用户姓名）
                row_user_name = str(row[0]).strip() if row[0] else ""
                if user_name and row_user_name != user_name:
                    continue

                direction = str(row[6]).strip() if row[6] else ""
                amount = row[7] or 0
                category = str(row[9]).strip() if row[9] else "未知"
                counter = str(row[11]).strip() if row[11] else "未知"
                date_str = str(row[2]).strip() if row[2] else ""

                key = f"{year:04d}-{month:02d}"

                if direction == "收入":
                    monthly_income[key] += amount
                    income_sources[counter] += amount
                    monthly_income_detail[key].append({"counter": counter, "amount": amount})

                    # 判断是否为个人转入
                    if _is_personal_transfer(category, counter):
                        personal_transfers.append({
                            "date": date_str,
                            "month": key,
                            "category": category,
                            "counterparty": counter,
                            "amount": amount,
                        })
                    else:
                        real_monthly_income[key] += amount

                elif direction == "支出":
                    monthly_expense[key] += amount
                    expense_categories[category] += amount
                    all_expenses.append({
                        "date": date_str,
                        "month": key,
                        "amount": amount,
                        "category": category,
                        "counterparty": counter,
                    })

            result["monthly_income"] = dict(monthly_income)
            result["monthly_expense"] = dict(monthly_expense)
            result["real_monthly_income"] = dict(real_monthly_income)
            result["income_sources"] = [
                {"name": name, "amount": round(amt, 0), "pct": round(amt / sum(income_sources.values()) * 100, 1)}
                for name, amt in income_sources.most_common(10)
            ]
            result["expense_categories"] = [
                {"name": name, "amount": round(amt, 0), "pct": round(amt / sum(expense_categories.values()) * 100, 1)}
                for name, amt in expense_categories.most_common(10)
            ]
            result["anomalies"] = self._identify_anomalies(monthly_income_detail)

            # 整理个人转入数据
            result["personal_transfers"] = sorted(personal_transfers, key=lambda x: x["date"])

            # 推断转入资金用途
            if personal_transfers:
                result["transfer_inference"] = self._infer_transfer_purpose(
                    personal_transfers, all_expenses
                )
                result["transfer_summary"] = self._summarize_transfers(personal_transfers)

        # 解析信用卡账单 (保持不变)
        if credit_card_file.exists():
            wb = openpyxl.load_workbook(credit_card_file)
            ws = wb.active

            cc_monthly = defaultdict(lambda: {"expense": 0, "count": 0})
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    year = int(row[2])
                    month = int(row[3])
                except (ValueError, TypeError):
                    continue

                # 信用卡按持卡人姓名过滤
                card_holder = str(row[5]).strip() if row[5] else ""
                if user_name and card_holder != user_name:
                    continue

                direction = str(row[11]).strip() if row[11] else ""
                amount = row[16] or 0

                key = f"{year:04d}-{month:02d}"
                if direction == "支出":
                    cc_monthly[key]["expense"] += amount
                    cc_monthly[key]["count"] += 1

            # 按年汇总
            cc_yearly = defaultdict(lambda: {"expense": 0, "count": 0, "months": 0})
            for m, v in sorted(cc_monthly.items()):
                y = m[:4]
                cc_yearly[y]["expense"] += v["expense"]
                cc_yearly[y]["count"] += v["count"]
                cc_yearly[y]["months"] += 1

            result["credit_card_trend"] = {
                y: {
                    "total": round(v["expense"], 0),
                    "monthly_avg": round(v["expense"] / max(v["months"], 1), 0),
                    "avg_per_txn": round(v["expense"] / max(v["count"], 1), 0),
                    "count": v["count"],
                }
                for y, v in sorted(cc_yearly.items())
            }

        return result

    def _infer_transfer_purpose(self, transfers: list[dict], expenses: list[dict]) -> dict:
        """根据转入后的消费行为，推断个人转入资金的用途

        规则：对每笔大额转入（>5000元），查找转入后30天内的消费，
        匹配用途关键词来判断资金去向。
        """
        if not transfers or not expenses:
            return {"inferred_purposes": [], "note": "数据不足，无法推断"}

        inferences = []
        for transfer in transfers:
            amount = transfer["amount"]
            # 仅对大额转入（>5000）做推断
            if amount < 5000:
                continue

            transfer_date = transfer["date"]
            try:
                td = datetime.strptime(transfer_date, "%Y-%m-%d")
            except (ValueError, TypeError):
                # 尝试仅日期部分
                try:
                    td = datetime.strptime(transfer_date[:10], "%Y-%m-%d")
                except (ValueError, TypeError):
                    continue

            # 查找转入后30天内的支出
            window_end = td + timedelta(days=30)
            nearby_expenses = []
            for exp in expenses:
                try:
                    ed = datetime.strptime(exp["date"][:10], "%Y-%m-%d")
                except (ValueError, TypeError):
                    continue
                if td <= ed <= window_end:
                    nearby_expenses.append(exp)

            if not nearby_expenses:
                inferences.append({
                    "date": transfer_date,
                    "amount": amount,
                    "counterparty": transfer["counterparty"],
                    "purpose": "unknown",
                    "purpose_label": "用途未明",
                    "detail": "转入后30天内无明显消费，可能为账户间资金调拨",
                })
                continue

            # 匹配用途规则
            matched_purpose = None
            max_matched = 0
            nearby_expense_categories = Counter(e["category"] for e in nearby_expenses)
            nearby_total = sum(e["amount"] for e in nearby_expenses)

            for keywords, purpose, label in TRANSFER_PURPOSE_RULES:
                # 统计匹配的支出金额
                matched_amount = 0
                for exp in nearby_expenses:
                    combined = exp["category"] + exp["counterparty"]
                    if any(kw in combined for kw in keywords):
                        matched_amount += exp["amount"]
                if matched_amount > max_matched:
                    max_matched = matched_amount
                    matched_purpose = (purpose, label, matched_amount)

            if matched_purpose and max_matched > 0:
                purpose, label, matched_amount = matched_purpose
                confidence = "高" if max_matched / amount > 0.3 else "中"
                inferences.append({
                    "date": transfer_date,
                    "amount": amount,
                    "counterparty": transfer["counterparty"],
                    "purpose": purpose,
                    "purpose_label": label,
                    "detail": f"转入¥{amount:,.0f}，后续30天内{label}类消费¥{matched_amount:,.0f}，置信度：{confidence}",
                })
            else:
                # 无匹配关键词，按最近消费分类推断
                top_cats = nearby_expense_categories.most_common(2)
                top_cat_labels = [c for c, _ in top_cats]
                inferences.append({
                    "date": transfer_date,
                    "amount": amount,
                    "counterparty": transfer["counterparty"],
                    "purpose": "general_cashflow",
                    "purpose_label": "一般资金周转",
                    "detail": f"转入¥{amount:,.0f}，30天内主要消费为{', '.join(top_cat_labels)}（共¥{nearby_total:,.0f}）",
                })

        return {
            "inferred_purposes": inferences,
            "total_inferred": len(inferences),
        }

    def _summarize_transfers(self, transfers: list[dict]) -> dict:
        """汇总个人转入数据"""
        total_amount = sum(t["amount"] for t in transfers)
        by_counterparty = defaultdict(lambda: {"count": 0, "total": 0.0})
        by_month = defaultdict(float)
        by_category = Counter()

        for t in transfers:
            cp = t["counterparty"]
            by_counterparty[cp]["count"] += 1
            by_counterparty[cp]["total"] += t["amount"]
            by_month[t["month"]] += t["amount"]
            by_category[t["category"]] += 1

        return {
            "total_count": len(transfers),
            "total_amount": round(total_amount, 2),
            "by_counterparty": [
                {"name": cp, "count": v["count"], "total": round(v["total"], 2)}
                for cp, v in sorted(by_counterparty.items(), key=lambda x: x[1]["total"], reverse=True)
            ],
            "by_month": {m: round(v, 2) for m, v in sorted(by_month.items())},
            "by_category": dict(by_category.most_common()),
        }

    def _identify_anomalies(self, monthly_income_detail: dict) -> list:
        """识别异常收入月份（奖金/年终奖等）"""
        monthly_totals = []
        for month, details in monthly_income_detail.items():
            total = sum(d["amount"] for d in details)
            monthly_totals.append((month, total))

        if len(monthly_totals) < 6:
            return []

        totals = [v for _, v in monthly_totals]
        avg = sum(totals) / len(totals)
        std = (sum((x - avg) ** 2 for x in totals) / len(totals)) ** 0.5

        anomalies = []
        for month, total in monthly_totals:
            if total > avg + 2 * std:
                anomalies.append({
                    "month": month,
                    "amount": round(total, 0),
                    "type": "收入异常高（可能是奖金/年终奖）",
                    "avg": round(avg, 0),
                })
            elif total < avg - std and total > 0:
                anomalies.append({
                    "month": month,
                    "amount": round(total, 0),
                    "type": "收入异常低",
                    "avg": round(avg, 0),
                })

        return sorted(anomalies, key=lambda x: x["month"])[-6:]

    def _compute_income_trend(self, monthly_income: dict) -> dict:
        """计算收入趋势指标"""
        if not monthly_income:
            return {}

        sorted_months = sorted(monthly_income.keys())
        values = [monthly_income[m] for m in sorted_months]

        # 按年汇总
        yearly = defaultdict(float)
        year_months = defaultdict(int)
        for m in sorted_months:
            y = m[:4]
            yearly[y] += monthly_income[m]
            year_months[y] += 1

        years = sorted(yearly.keys())
        annualized = {}
        for y in years:
            annualized[y] = yearly[y] / max(year_months[y], 1) * 12

        # 计算增长率
        growth_rates = {}
        for i in range(1, len(years)):
            prev, curr = years[i - 1], years[i]
            if annualized[prev] > 0:
                growth_rates[curr] = round((annualized[curr] / annualized[prev] - 1) * 100, 1)

        # 稳定性分析
        avg_monthly = sum(values) / len(values)
        std = (sum((x - avg_monthly) ** 2 for x in values) / len(values)) ** 0.5
        cv = std / avg_monthly if avg_monthly > 0 else 0  # 变异系数

        # 趋势方向（最近6个月 vs 前6个月）
        recent = values[-6:] if len(values) >= 6 else values
        earlier = values[-12:-6] if len(values) >= 12 else values[:max(1, len(values) // 2)]
        recent_avg = sum(recent) / len(recent) if recent else 0
        earlier_avg = sum(earlier) / len(earlier) if earlier else 0
        short_trend = round((recent_avg / earlier_avg - 1) * 100, 1) if earlier_avg > 0 else 0

        return {
            "total_months": len(sorted_months),
            "yearly_income": {y: round(v, 0) for y, v in sorted(yearly.items())},
            "annualized_income": {y: round(v, 0) for y, v in sorted(annualized.items())},
            "growth_rates": growth_rates,
            "avg_monthly": round(avg_monthly, 0),
            "stability": "稳定" if cv < 0.3 else "波动" if cv < 0.6 else "剧烈波动",
            "cv": round(cv, 2),
            "short_trend_pct": short_trend,
            "short_trend_direction": "上升" if short_trend > 5 else "下降" if short_trend < -5 else "平稳",
        }

    def _compute_expense_trend(self, monthly_expense: dict) -> dict:
        """计算支出趋势指标"""
        if not monthly_expense:
            return {}

        sorted_months = sorted(monthly_expense.keys())
        values = [monthly_expense[m] for m in sorted_months]

        yearly = defaultdict(float)
        year_months = defaultdict(int)
        for m in sorted_months:
            y = m[:4]
            yearly[y] += monthly_expense[m]
            year_months[y] += 1

        years = sorted(yearly.keys())
        annualized = {y: yearly[y] / max(year_months[y], 1) * 12 for y in years}

        growth_rates = {}
        for i in range(1, len(years)):
            prev, curr = years[i - 1], years[i]
            if annualized[prev] > 0:
                growth_rates[curr] = round((annualized[curr] / annualized[prev] - 1) * 100, 1)

        avg_monthly = sum(values) / len(values)
        std = (sum((x - avg_monthly) ** 2 for x in values) / len(values)) ** 0.5
        cv = std / avg_monthly if avg_monthly > 0 else 0

        recent = values[-6:] if len(values) >= 6 else values
        earlier = values[-12:-6] if len(values) >= 12 else values[:max(1, len(values) // 2)]
        recent_avg = sum(recent) / len(recent) if recent else 0
        earlier_avg = sum(earlier) / len(earlier) if earlier else 0
        short_trend = round((recent_avg / earlier_avg - 1) * 100, 1) if earlier_avg > 0 else 0

        return {
            "total_months": len(sorted_months),
            "yearly_expense": {y: round(v, 0) for y, v in sorted(yearly.items())},
            "annualized_expense": {y: round(v, 0) for y, v in sorted(annualized.items())},
            "growth_rates": growth_rates,
            "avg_monthly": round(avg_monthly, 0),
            "stability": "稳定" if cv < 0.3 else "波动" if cv < 0.6 else "剧烈波动",
            "cv": round(cv, 2),
            "short_trend_pct": short_trend,
            "short_trend_direction": "上升" if short_trend > 5 else "下降" if short_trend < -5 else "平稳",
        }

    def _forecast_balance(self, income_trend: dict, expense_trend: dict) -> dict:
        """预测未来收支平衡"""
        if not income_trend or not expense_trend:
            return {}

        # 获取最新年收入和支出（年化）
        inc_annual = income_trend.get("annualized_income", {})
        exp_annual = expense_trend.get("annualized_expense", {})

        if not inc_annual or not exp_annual:
            return {}

        latest_year = max(set(inc_annual.keys()) & set(exp_annual.keys()), default=None)
        if latest_year is None:
            return {}

        latest_income = inc_annual[latest_year]
        latest_expense = exp_annual[latest_year]

        # 计算平均增长率
        inc_growth = income_trend.get("growth_rates", {})
        exp_growth = expense_trend.get("growth_rates", {})

        avg_inc_growth = sum(inc_growth.values()) / len(inc_growth) if inc_growth else 0
        avg_exp_growth = sum(exp_growth.values()) / len(exp_growth) if exp_growth else 0

        # 预测未来2年
        forecast_years = 2
        forecast = []
        current_income = latest_income
        current_expense = latest_expense

        for year in range(1, forecast_years + 1):
            current_income *= (1 + avg_inc_growth / 100)
            current_expense *= (1 + avg_exp_growth / 100)
            balance = current_income - current_expense
            balance_rate = balance / current_income * 100 if current_income > 0 else 0

            forecast.append({
                "year": f"第{year}年",
                "income": round(current_income, 0),
                "expense": round(current_expense, 0),
                "balance": round(balance, 0),
                "balance_rate": round(balance_rate, 1),
            })

        # 判断能否覆盖
        can_cover = avg_inc_growth >= avg_exp_growth
        gap_widening = avg_exp_growth > avg_inc_growth

        return {
            "forecast_years": forecast_years,
            "latest_year": latest_year,
            "latest_income": round(latest_income, 0),
            "latest_expense": round(latest_expense, 0),
            "latest_balance": round(latest_income - latest_expense, 0),
            "latest_balance_rate": round((latest_income - latest_expense) / latest_income * 100, 1) if latest_income > 0 else 0,
            "avg_income_growth": round(avg_inc_growth, 1),
            "avg_expense_growth": round(avg_exp_growth, 1),
            "forecast": forecast,
            "can_cover": can_cover,
            "gap_widening": gap_widening,
            "risk_level": "低风险" if can_cover else "中风险" if not gap_widening else "高风险",
            "conclusion": (
                "收入增速可覆盖消费增长，结余率有望改善"
                if can_cover
                else "收入增速低于消费增速，收支缺口可能扩大，建议调整消费结构"
            ),
        }

    def _generate_strategies(self, income_trend: dict, expense_trend: dict, forecast: dict, balance: float) -> list:
        """生成消费调整策略"""
        strategies = []

        # 策略1: 应急储备
        avg_monthly_expense = expense_trend.get("avg_monthly", 0)
        months_covered = balance / avg_monthly_expense if avg_monthly_expense > 0 else 0

        if months_covered < 6:
            strategies.append({
                "priority": "P0",
                "category": "风险缓冲",
                "title": "建立应急储备金",
                "description": f"当前余额可支撑约{months_covered:.1f}个月支出，建议储备6个月生活费（约¥{avg_monthly_expense * 6:,.0f}）",
                "action": "每月固定存入收入的10-15%作为应急基金",
                "impact": "高",
            })

        # 策略2: 收入端
        inc_stability = income_trend.get("stability", "")
        if inc_stability in ["波动", "剧烈波动"]:
            strategies.append({
                "priority": "P0",
                "category": "收入端",
                "title": "稳定收入来源",
                "description": f"收入变异系数为{income_trend.get('cv', 0)}，存在较大波动，建议关注主业稳定性",
                "action": "评估职场风险，考虑副业或技能提升以增加收入稳定性",
                "impact": "高",
            })

        # 策略3: 支出端 - 根据消费趋势
        exp_direction = expense_trend.get("short_trend_direction", "")
        exp_growth = expense_trend.get("growth_rates", {})
        if exp_direction == "上升":
            strategies.append({
                "priority": "P1",
                "category": "支出端",
                "title": "控制消费增长",
                "description": "近期消费呈上升趋势，建议审视非必要支出",
                "action": "区分刚性支出（房贷/医疗/教育）与弹性支出（餐饮/购物/娱乐），优先压缩弹性支出20%",
                "impact": "中",
            })

        # 策略4: 结构性优化
        if forecast.get("gap_widening", False):
            strategies.append({
                "priority": "P1",
                "category": "结构性调整",
                "title": "量入为出，优化支出结构",
                "description": forecast.get("conclusion", ""),
                "action": "制定月度预算，将支出控制在收入的80%以内，结余用于储蓄和投资",
                "impact": "高",
            })

        # 策略5: 信用卡管理
        strategies.append({
            "priority": "P2",
            "category": "负债管理",
            "title": "优化信用卡使用",
            "description": "避免信用卡分期和高额利息支出",
            "action": "全额还款，减少分期；利用免息期优化现金流",
            "impact": "中",
        })

        # 策略6: 长期规划
        strategies.append({
            "priority": "P2",
            "category": "长期规划",
            "title": "建立定期储蓄习惯",
            "description": "无论收入波动如何，保持固定储蓄比例",
            "action": "设置自动转账，每月工资到账后立即转出固定比例（建议15-20%）到储蓄账户",
            "impact": "中",
        })

        return sorted(strategies, key=lambda x: x["priority"])
