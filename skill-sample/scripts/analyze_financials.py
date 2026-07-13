#!/usr/bin/env python3
"""
财务状况分析器 — 收入 vs 消费增长趋势分析
从交易流水和信用卡账单中提取收支数据，分析财务健康度和增长趋势。
"""
import json
import sys
from pathlib import Path
from collections import defaultdict, Counter

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)


def parse_transactions(filepath):
    """解析交易流水，返回月度收支数据"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    monthly = defaultdict(lambda: {'income': 0, 'expense': 0, 'income_count': 0, 'expense_count': 0})
    income_sources = Counter()
    expense_categories = Counter()
    monthly_income_sources = defaultdict(lambda: Counter())
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        try:
            year = int(row[2])
            month = int(row[3])
        except (ValueError, TypeError):
            continue
        
        direction = str(row[5]).strip()
        amount = row[6] or 0
        category = str(row[8]).strip() if row[8] else '未知'
        counter = str(row[10]).strip() if row[10] else '未知'
        
        key = f'{year:04d}-{month:02d}'
        
        if direction == '收入':
            monthly[key]['income'] += amount
            monthly[key]['income_count'] += 1
            income_sources[counter] += amount
            monthly_income_sources[key][counter] += amount
        elif direction == '支出':
            monthly[key]['expense'] += amount
            monthly[key]['expense_count'] += 1
            expense_categories[category] += amount
    
    return monthly, income_sources, expense_categories, monthly_income_sources


def parse_credit_card(filepath):
    """解析信用卡账单"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    monthly = defaultdict(lambda: {'expense': 0, 'count': 0})
    category_trend = defaultdict(lambda: defaultdict(float))
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            year = int(row[2])
            month = int(row[3])
        except (ValueError, TypeError):
            continue
        
        direction = str(row[11]).strip() if row[11] else ''
        amount = row[16] or 0
        cat = str(row[15]).strip() if row[15] else '未知'
        
        key = f'{year:04d}-{month:02d}'
        
        if direction == '支出':
            monthly[key]['expense'] += amount
            monthly[key]['count'] += 1
            category_trend[cat][key] += amount
    
    return monthly, category_trend


def compute_growth_rates(monthly):
    """计算年度增长率和趋势"""
    sorted_months = sorted(monthly.keys())
    
    yearly = defaultdict(lambda: {'income': 0, 'expense': 0, 'months': 0})
    for m in sorted_months:
        y = m[:4]
        yearly[y]['income'] += monthly[m]['income']
        yearly[y]['expense'] += monthly[m]['expense']
        yearly[y]['months'] += 1
    
    years = sorted(yearly.keys())
    growth = {}
    
    for i in range(1, len(years)):
        prev_y, curr_y = years[i-1], years[i]
        prev_inc = yearly[prev_y]['income']
        curr_inc = yearly[curr_y]['income']
        prev_exp = yearly[prev_y]['expense']
        curr_exp = yearly[curr_y]['expense']
        
        # Annualize partial years
        prev_months = yearly[prev_y]['months']
        curr_months = yearly[curr_y]['months']
        prev_inc_annual = prev_inc / prev_months * 12
        curr_inc_annual = curr_inc / curr_months * 12
        prev_exp_annual = prev_exp / prev_months * 12
        curr_exp_annual = curr_exp / curr_months * 12
        
        growth[curr_y] = {
            'income_growth': round((curr_inc_annual / prev_inc_annual - 1) * 100, 1),
            'expense_growth': round((curr_exp_annual / prev_exp_annual - 1) * 100, 1),
            'monthly_avg_income': round(curr_inc / curr_months, 0),
            'monthly_avg_expense': round(curr_exp / curr_months, 0),
            'annualized_income': round(curr_inc_annual, 0),
            'annualized_expense': round(curr_exp_annual, 0),
        }
    
    return yearly, growth


def compute_monthly_trend(monthly):
    """计算月度趋势"""
    sorted_months = sorted(monthly.keys())
    trend = []
    for m in sorted_months:
        d = monthly[m]
        trend.append({
            'month': m,
            'income': round(d['income'], 2),
            'expense': round(d['expense'], 2),
            'balance': round(d['income'] - d['expense'], 2),
            'balance_rate': round((d['income'] - d['expense']) / d['income'] * 100, 1) if d['income'] > 0 else None
        })
    return trend


def identify_anomalies(monthly):
    """识别异常月份"""
    sorted_data = []
    for m in sorted(monthly.keys()):
        d = monthly[m]
        sorted_data.append((m, d['income']))
    
    if len(sorted_data) < 6:
        return []
    
    incomes = [v for _, v in sorted_data]
    avg = sum(incomes) / len(incomes)
    std = (sum((x - avg)**2 for x in incomes) / len(incomes)) ** 0.5
    
    anomalies = []
    for m, v in sorted_data:
        if v > avg + 2 * std:
            anomalies.append({'month': m, 'amount': round(v, 0), 'type': '收入异常高（可能是奖金/年终）'})
        elif v < avg - std and v > 0:
            anomalies.append({'month': m, 'amount': round(v, 0), 'type': '收入异常低'})
    
    return anomalies[-6:]  # 最近6个


def main():
    base_dir = Path(__file__).parent.parent.parent.parent.parent / 'coredatas'
    
    transactions_file = base_dir / '银行交易流水汇总明细表.xlsx'
    credit_card_file = base_dir / '信用卡账单合并明细表.xlsx'
    accounts_file = base_dir / '客户账户信息表.xlsx'
    
    result = {
        'monthly_trend': [],
        'yearly_summary': {},
        'growth_rates': {},
        'income_sources': [],
        'expense_categories': [],
        'credit_card_trend': {},
        'category_trend': {},
        'anomalies': [],
        'account_summary': {},
    }
    
    # 分析交易流水
    if transactions_file.exists():
        monthly, income_sources, expense_categories, _ = parse_transactions(transactions_file)
        
        result['monthly_trend'] = compute_monthly_trend(monthly)
        
        yearly, growth = compute_growth_rates(monthly)
        result['yearly_summary'] = {
            y: {
                'income': round(v['income'], 0),
                'expense': round(v['expense'], 0),
                'balance': round(v['income'] - v['expense'], 0),
                'balance_rate': round((v['income'] - v['expense']) / v['income'] * 100, 1) if v['income'] > 0 else 0,
                'months': v['months']
            }
            for y, v in sorted(yearly.items())
        }
        result['growth_rates'] = growth
        
        result['income_sources'] = [
            {'name': name, 'amount': round(amt, 0), 'pct': round(amt / sum(v for _, v in income_sources.most_common()) * 100, 1)}
            for name, amt in income_sources.most_common(10)
        ]
        
        result['expense_categories'] = [
            {'name': name, 'amount': round(amt, 0), 'pct': round(amt / sum(v for _, v in expense_categories.most_common()) * 100, 1)}
            for name, amt in expense_categories.most_common(10)
        ]
        
        result['anomalies'] = identify_anomalies(monthly)
    
    # 分析信用卡账单
    if credit_card_file.exists():
        cc_monthly, category_trend = parse_credit_card(credit_card_file)
        
        cc_sorted = sorted(cc_monthly.keys())
        cc_yearly = defaultdict(lambda: {'expense': 0, 'count': 0, 'months': 0})
        for m in cc_sorted:
            y = m[:4]
            cc_yearly[y]['expense'] += cc_monthly[m]['expense']
            cc_yearly[y]['count'] += cc_monthly[m]['count']
            cc_yearly[y]['months'] += 1
        
        result['credit_card_trend'] = {
            y: {
                'total_expense': round(v['expense'], 0),
                'monthly_avg': round(v['expense'] / v['months'], 0),
                'avg_per_transaction': round(v['expense'] / v['count'], 0) if v['count'] else 0,
                'total_count': v['count']
            }
            for y, v in sorted(cc_yearly.items())
        }
        
        # 年度分类趋势
        result['category_trend'] = {}
        for cat in ['餐饮美食', '公共交通', '通讯软件数码', '超市便利店', '网约车', '医疗健康', '网购电商', '日用百货']:
            if cat in category_trend:
                yearly_cat = defaultdict(float)
                for m, amt in category_trend[cat].items():
                    yearly_cat[m[:4]] += amt
                result['category_trend'][cat] = {y: round(amt, 0) for y, amt in sorted(yearly_cat.items())}
    
    # 账户摘要
    if accounts_file.exists():
        wb = openpyxl.load_workbook(accounts_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=4, values_only=True):
            name = row[1]
            acc_type = row[2]
            balance = row[8]
            if name and acc_type:
                result['account_summary'][str(acc_type)] = round(balance, 2) if balance else 0
    
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
